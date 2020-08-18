
import rdkit
import logging
import tempfile
import time,datetime
from pathlib import Path
from argparse import ArgumentParser
from logging import getLogger
import shutil
from tarfile import TarFile
from configparser import ConfigParser
import pickle
import numpy as np
from tqdm import tqdm

import matplotlib as mpl
mpl.use('Agg')
from matplotlib import pyplot as plt
mpl.interactive(False) # modify memory leak

import rdkit
from rdkit import Chem
from rdkit.Chem import Draw
from rdkit.Chem import AllChem
from rdkit import DataStructs

import multiprocessing
from multiprocessing.pool import Pool

import torch

from torch_jtnn import *
from pcemg.model.ms_encoder import ms_peak_encoder_cnn as ms_peak_encoder_cnn

class Sampler():
    """ Sampling from dataset.
    
    Encoding in batches, decoding in data.
    
    Args:
        vali_dataset ():
        enc_model (nn.Module):
        dec_model (nn.Module):
        device (str):
    
    Yields:
        
    
    """
    def __init__(self,vali_dataset,enc_model,dec_model,device='cuda'):
        self.vali_dataset = vali_dataset
        self.__num_batch = len(vali_dataset)
        self.__batch_size = vali_dataset.batch_size
        self.enc_model = enc_model
        self.dec_model = dec_model
        
        self.device = device
        
    def __len__(self):
        self.__num_batch = len(self.vali_dataset)
        self.__batch_size = self.vali_dataset.batch_size
        return self.__num_batch * self.__batch_size
    
    def __iter__(self):
        
        vali_dataset = self.vali_dataset
        enc_model = self.enc_model
        dec_model = self.dec_model
        device = self.device
        
        for batch in vali_dataset:
            x_batch, x_jtenc_holder, x_mpn_holder, x_jtmpn_holder,x,y = batch
            gx = x.to(device)
            gy = y.to(device)
            gy.requires_grad_(True)
            
            h,_ = enc_model(gx,gy,sample=False)

            tree_vec = h[:,:int(h.shape[1]/2)]
            mol_vec  = h[:,int(h.shape[1]/2):]
            batch_num,latent_num = h.shape
            
            for b in range(batch_num):
                predict_smiles = dec_model.decode(tree_vec[b:b+1],mol_vec[b:b+1],False)
                pred_nodes = dec_model.pred_nodes
                pred_scores = dec_model.pred_scores
                
                g = torch.cat([torch.autograd.grad(h[b,n],enc_model.inp_,retain_graph=True)[0][b:b+1] for n in range(latent_num)])
                gnodes = torch.cat([torch.autograd.grad(score,enc_model.inp_,retain_graph=True)[0][b:b+1] for score in pred_scores])
                
                I = torch.cat([torch.autograd.grad(h[b,n],gy,retain_graph=True)[0][b:b+1] for n in range(latent_num)])
                Inodes = torch.cat([torch.autograd.grad(score,gy,retain_graph=True)[0][b:b+1] for score in pred_scores])
                
                res_dict = {
                    'input_mz':x.numpy()[b,:],
                    'input_I':y.numpy()[b,:],
                    'Saliency_z_mz':g.cpu().numpy(),
                    'Saliency_node_mz':gnodes.cpu().numpy(),
                    'Saliency_z_I':I.cpu().numpy(),
                    'Saliency_node_I':Inodes.cpu().numpy(),
                    'true_smiles':x_batch[b].smiles,
                    'pred_smiles':predict_smiles,
                    'pred_nodes_smiles':[node.smiles for node in pred_nodes],
                    'pred_scores':torch.stack(pred_scores).cpu().detach().numpy(),
                }
                yield res_dict
                #yield x.numpy()[b,:], y.numpy()[b,:], g.cpu().numpy(), gnodes.cpu().numpy(), x_batch[b].smiles, predict_smiles, pred_nodes


class SpectrumPloter():
    def __init__(self):
        pass
    
    def plot_spec_node(self,mz,inten,gr,psmiles,pscores,name):

        # write molecular structures
        fig = plt.figure(dpi=250)
        ax1 = fig.add_subplot(2,1,1)
        ax1.axis('off')
        img = Draw.MolsToGridImage([Chem.MolFromSmiles(psmiles)],molsPerRow=3)
        ax1.imshow(img)
        ax1.text(0.2,ax1.get_ylim()[1],"pred:{}, scores:{}".format(psmiles,pscores),size=5)

        # write a spectrum 
        self.__plot_spectrum(fig,mz,inten,gr)

        fig.savefig(name)
        plt.cla()
        plt.clf()
        plt.close()
    
    def plot_spec(self,mz,inten,gr,smiles,psmiles,name):

        # write molecular structures
        fig = plt.figure(dpi=250)
        ax1 = fig.add_subplot(2,1,1)
        ax1.axis('off')
        img = Draw.MolsToGridImage([Chem.MolFromSmiles(s) for s in [smiles,psmiles] if s is not None],molsPerRow=3)
        ax1.imshow(img)
        ax1.text(0.2,ax1.get_ylim()[1],"Real:{} - pred:{}".format(smiles,psmiles),size=5)

        # write a spectrum 
        self.__plot_spectrum(fig,mz,inten,gr)

        fig.savefig(name)
        plt.cla()
        plt.clf()
        plt.close()
    
    def __plot_spectrum(self,fig,mz,inten,gr):
        # write a spectrum 
        ax2 = fig.add_subplot(2,1,2)
        cmap = plt.get_cmap('cool')
        bar = ax2.bar(mz,inten,color=cmap(gr/np.max(gr)))
        sc = ax2.scatter(mz,inten,c=gr/np.max(gr),cmap=cmap,s=0)

        cbar = fig.colorbar(sc)
        cbar.set_label("rate of relevance [-]")

        ax2.set_xlabel('m/z')
        ax2.set_ylabel('int.')

class _proc():

    fingers = {
        'MACCS':AllChem.GetMACCSKeysFingerprint,
        'MORGAN':AllChem.GetMorganFingerprint,
    }

    def __init__(self,ploter,path,pre_func,fingerkey=('MACCS',)):
        self.ploter = ploter
        self.path = path
        self.pre_func = pre_func
        self.fingerkey = fingerkey
        
    def __call__(self,inputs):

        ploter = self.ploter
        path = self.path
        pre_func = self.pre_func
        fingerkey = self.fingerkey

        i,batch = inputs

        x = batch['input_mz']
        y = batch['input_I']
        g = batch['Saliency_z_mz']
        gI = batch['Saliency_z_I']
        gnodes = batch['Saliency_node_mz']
        gInodes = batch['Saliency_node_I']
        smiles = batch['true_smiles']
        psmiles = batch['pred_smiles']
        pnodes = batch['pred_nodes_smiles']
        pscores = batch['pred_scores']

        datapath = path / "data{no:03d}".format(no=i)
        datapath.mkdir(exist_ok=True)
        if pre_func == 'abs':
            g = np.abs(g).mean(axis=2)
        elif pre_func == 'l2':
            g = np.linalg.norm(g,axis=2)
        g = g.sum(axis=0)
        gI = gI.sum(axis=0)

        if pre_func == 'abs':
            gnodes= np.abs(gnodes).mean(axis=2)
        elif pre_func == 'l2':
            gnodes = np.linalg.norm(gnodes,axis=2)

        ploter.plot_spec(x,y,g,smiles,psmiles,str(datapath.joinpath("Saliency_mz.png")).format(no=i))
        ploter.plot_spec(x,y,gI,smiles,psmiles,str(datapath.joinpath("Saliency_I.png")).format(no=i))
        for j,(_gnode,_gInode,pnode,pscore) in enumerate(zip(gnodes,gInodes,pnodes,pscores)):
            node_path = datapath / "nodes"
            node_path.mkdir(exist_ok=True)
            ploter.plot_spec_node(x,y,_gnode,pnode,pscore,str(node_path.joinpath("Saliency_mz_{cl:03d}.png")).format(cl=j))
            ploter.plot_spec_node(x,y,_gInode,pnode,pscore,str(node_path.joinpath("Saliency_I_{cl:03d}.png")).format(cl=j))
        
        batch['similarity'] = self.calcSimilarity(smiles,psmiles,key=fingerkey)

        return batch

    def calcSimilarity(self,tsmiles,psmiles,key=('MACCS',)):
        tmol = Chem.MolFromSmiles(tsmiles)
        pmol = Chem.MolFromSmiles(psmiles)
        finger = self.fingers[key[0]]
        tfinger = finger(tmol,*key[1:])
        pfinger = finger(pmol,*key[1:])

        return DataStructs.TanimotoSimilarity(tfinger,pfinger)

class SaliencyCalc(object):
    def __init__(self,trained_result_path,eval_mode,logger=None):
        self.__trained_result_path = trained_result_path
        self.__eval_mode  = eval_mode
        self.__logger = logger
        self.starttime = datetime.datetime.fromtimestamp(time.time())

        lg = rdkit.RDLogger.logger() 
        lg.setLevel(rdkit.RDLogger.CRITICAL)

    def __call__(self):
        try:
            tempdir = tempfile.TemporaryDirectory(prefix='analysis_model-')
            temp_path = Path(tempdir.name)

            self.__logger.info(f"Temp path:{temp_path}")
            log = (temp_path/"log").open('w')

            self.copy_require_file(tempdir)
            config = self.load_config()

            with open(self.vali_file,'rb') as f:
                vali_dataset = pickle.load(f)
            vali_dataset.shuffle = False

            with open(self.vocab_path,'r') as f:
                vocab,_ = zip(*[(x.split(',')[0].strip('\n\r'),int(x.split(',')[1].strip('\n\r'))) for x in f ])
                
            vocab = Vocab(vocab,_)

            dec_model = JTNNVAE(vocab=vocab,**config['JTVAE']).to('cuda')
            
            enc_model = ms_peak_encoder_cnn(vali_dataset.max_spectrum_size,**config['PEAK_ENCODER'],varbose=False).to('cuda')
            
            print(f"Structuer of Encoder\n {enc_model}")
            print(f"Structuer of Decoder\n {dec_model}")
            enc_model_path,dec_model_path = SaliencyCalc.select_iter(self.model_list,self.iter_list)
            self.__logger.info(f"{enc_model_path},{dec_model_path}")
            

            enc_model.load_state_dict(torch.load(enc_model_path,map_location='cuda'))
            dec_model.load_state_dict(torch.load(dec_model_path,map_location='cuda'))

            self.VanillaGrad(enc_model,dec_model,vali_dataset)

        finally:
            pass
        pass

    def VanillaGrad(self,enc_model,dec_model,vali_dataset,device='cuda',pre_func='abs',result=None):
        vali_dataset.batch_size = 10
        vali_dataset.shuffle = False
        assert pre_func in ('abs','l2')
        
        result = result or "./result"
        path = Path(result)
        path.mkdir(exist_ok=True)
        ploter = SpectrumPloter()
        
        processed = Path("saliency.pkl")
        if not processed.exists():
            with Pool() as p:
                pool_iter = p.imap(_proc(ploter,path,pre_func,fingerkey=("MORGAN",2)),enumerate(Sampler(vali_dataset,enc_model,dec_model,device='cuda')))
                results = list(tqdm(pool_iter,total=len(vali_dataset)*vali_dataset.batch_size))
            with processed.open('wb') as f:
                pickle.dump(results,f)           

        with processed.open('rb') as f:
            results = pickle.load(f)

        self.log_output(results)
    
    def log_output(self,results):
        from openpyxl import Workbook

        def _one_data(n,result,wb):
            sheet = wb.create_sheet(f"data{n:03d}")
            sheet.append(['true smiles','predict smiles'])
            sheet.append([result['true_smiles'],result['pred_smiles']])
            sheet.append(['similarity',result['similarity']])

            sheet.append(['spectrum peaks'])
            sheet.append(result['input_mz'].tolist())
            sheet.append(np.abs(result['Saliency_z_mz']).mean(axis=2).sum(axis=0).tolist())

            sheet.append([])
            sheet.append(result['pred_scores'].tolist())

        ndata = len(results)

        wb = Workbook()
        self.__logger.info("logging")

        similarites,pred_scores = zip(*[(result['similarity'],float(result['pred_scores'].mean())) for result in results])
        sheet = wb.create_sheet('total')
        sheet.append(similarites)
        sheet.append(pred_scores)


        for n,result in enumerate(results):
            _one_data(n,result,wb)

        
        wb.save("saliency_.xlsx")


    def copy_require_file(self,tempdir):
        temp_path = Path(tempdir.name)
        trained_result_path = shutil.copy(self.__trained_result_path,temp_path)

        with TarFile.open(trained_result_path,"r:gz") as f:
            f.extractall(path=temp_path)

        self.model_list,self.iter_list = self.find_trained_model(temp_path)
        self.model_config = self.find_config_file(temp_path)
        self.vocab_path = self.find_vocab_file(temp_path)
        self.vali_file = self.find_validata(temp_path)

    def load_config(self):
        with self.model_config.open('r') as f:
            print(f.read())

        config = ConfigParser()
        config.optionxform=str
        config.read(self.model_config)
        self.__logger.info(f"{config.sections()}")
        ret = dict()
        ret['JTVAE'] = JTNNVAE.config_dict(config['JTVAE'])
        #ret['PEAK_ENCODER'] = ms_peak_encoder_cnn.config_dict(config['PEAK_ENCODER'])
        ret['PEAK_ENCODER'] = config['PEAK_ENCODER']

        return ret

    def find_trained_model(self,temp_path):
        
        def extract_iter_num(mlist):
            for path in mlist:
                try:
                    yield int(path.name.split('-')[1])
                except ValueError:
                    pass
            #return int(path.name.split('-')[1])

        enc_model_list = list(temp_path.glob("./**/enc_model/model.iter-*"))
        dec_model_list = list(temp_path.glob("./**/dec_model/model.iter-*"))
        
        iter_list = sorted([n for n in extract_iter_num(enc_model_list)])

        model_list = {
            'encoder':enc_model_list,
            'decoder':dec_model_list,
        }
        return model_list,iter_list
    
    def find_config_file(self,temp_path):
        model_config = list(temp_path.glob('./**/model_config.ini'))[0]
        self.__logger.info(f"Model Configuration file path:{model_config}")
        return model_config
    
    def find_vocab_file(self,temp_path):
        vocab = list(temp_path.glob('./**/MS_vocab.txt'))[0]
        self.__logger.info(vocab)
        return vocab
        
    def find_validata(self,temp_path):
        vali = list(temp_path.glob('./**/vali_data.pkl'))[0]
        self.__logger.info(vali)
        return vali

    @staticmethod
    def select_iter(model_list,iter_list):
        if len(iter_list) > 5:
            header = "please choice iterate [" + str(iter_list[0]) + " " + str(iter_list[1]) + " ... " + str(iter_list[-2]) + " " + str(iter_list[-1])+"] >> "
        else:
            header = "please choise iterate ["
            for i in iter_list:
                header += str(i) + " "
            header += "] >> "

        select_iter = input(header)
        #if not int(select_iter) in iter_list:
        #    raise IndexError()

        enc_model = None
        dec_model = None

        for model_path in model_list['encoder']:
             if "model.iter-"+select_iter in model_path.name:
                 enc_model = model_path

        for model_path in model_list['decoder']:
             if "model.iter-"+select_iter in model_path.name:
                 dec_model = model_path

        assert enc_model is not None and dec_model is not None

        return enc_model,dec_model

def saliencyCalc(trained_result_path,eval_mode=True,logger=None,loglevel=logging.WARN):
    if logger is None:
        rv = logging._checkLevel(loglevel)
        logger = getLogger(__name__)
        stream_handler = logging.StreamHandler()
        stream_handler.setLevel(rv)
        stream_handler.setFormatter(logging.Formatter("[%(levelname)s]:%(message)s"))
        logger.addHandler(stream_handler)
        logger.setLevel(rv)

    SaliencyCalc(trained_result_path,eval_mode,logger=logger)()

def main():
    lg = rdkit.RDLogger.logger() 
    lg.setLevel(rdkit.RDLogger.CRITICAL)
    parser = ArgumentParser()

    parser.add_argument('trained_result_path',type=str)
    parser.add_argument('--loglevel',default='WARN')

    args = parser.parse_args()

    proc = SaliencyCalc(**args.__dict__)
    proc()

if __name__=="__main__":
    main()